#!/usr/bin/env python3
"""Preprocess sales XML + product catalog XLSX → JSON for React dashboard."""

import csv
import json
import re
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Optional, Set, Dict, List, Any

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
SALES_XML = BASE_DIR / "product - 2026-04-06T201032.237.xml"
CATALOG_XLSX = BASE_DIR / "Выгрузка_DE.xlsx"
LIEFERANT_XLSX = BASE_DIR / "SKU-Lieferant.xlsx"
INVENTORY_TXT = BASE_DIR / "Lagerbestand+f&uuml;r+Versand+durch+Amazon_04-06-2026.txt"
OUT_DIR = BASE_DIR / "dashboard" / "public" / "data"

NS = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}

# camelCase mapping for sales columns
SALES_KEY_MAP = {
    "Bestellung #": "bestellungNr",
    "Status des Auftrags": "status",
    "Bestelldatum": "bestelldatum",
    "Artikelposition": "artikelposition",
    "Kunden Email": "kundenEmail",
    "Kundenname": "kundenname",
    "Kundengruppe": "kundengruppe",
    "Land": "land",
    "Region": "region",
    "Stadt": "stadt",
    "Postleitzahl": "postleitzahl",
    "Adresse": "adresse",
    "Phone": "phone",
    "Produktbezeichnung": "produktbezeichnung",
    "Hersteller": "hersteller",
    "Qty. Ordered": "qtyOrdered",
    "Qty. Invoiced": "qtyInvoiced",
    "Qty. Shipped": "qtyShipped",
    "Qty. Refunded": "qtyRefunded",
    "Preis": "preis",
    "Originalpreis": "originalpreis",
    "Zwischensumme": "zwischensumme",
    "Discounts": "discounts",
    "MwSt.": "mwst",
    "Gesamt": "gesamt",
    "Total Incl. Tax": "totalInclTax",
    "In Rechnung gestellt.": "inRechnungGestellt",
    "Tax Invoiced": "taxInvoiced",
    "Invoiced Incl. Tax": "invoicedInclTax",
    "Rückerstattet": "rueckerstattet",
    "Tax Refunded": "taxRefunded",
    "Refunded Incl. Tax": "refundedInclTax",
    "Total Cost": "totalCost",
    "Total Revenue (excl.tax)": "totalRevenueExclTax",
    "Total Revenue": "totalRevenue",
    "Total Profit": "totalProfit",
    "Total Margin": "totalMargin",
}

MONEY_FIELDS = {
    "preis", "originalpreis", "zwischensumme", "discounts", "mwst",
    "gesamt", "totalInclTax", "inRechnungGestellt", "taxInvoiced",
    "invoicedInclTax", "rueckerstattet", "taxRefunded", "refundedInclTax",
    "totalCost", "totalRevenueExclTax", "totalRevenue", "totalProfit",
}

QTY_FIELDS = {"qtyOrdered", "qtyInvoiced", "qtyShipped", "qtyRefunded"}

PRODUCT_COLUMNS = [
    "sku", "sku_vender", "purchase_price", "amaz_parent_sku",
    "amaz_name", "chain_length_google", "price", "amaz_price",
    "status", "amaz_chain_type", "chain_metal_type", "chain_metal_aloy",
    "chain_type", "chain_length", "chain_width", "product_type",
    "amaz_metal_stamp", "chain_weight",
    "earring_weight", "earring_diameter", "earring_type", "earring_width",
    "earring_fassung", "pendant_weight", "pendant_width", "pendant_height",
    "pendant_type", "pendant_metal_aloy", "pendant_metal_type",
    "ringe_size", "ringe_metal_aloy", "ringe_fassung", "ringe_produkt_type",
]

LIEFERANT_ALIASES = {
    "top gold": "Top Gold",
}


def parse_money(val: Optional[str]) -> Optional[float]:
    if not val:
        return None

    cleaned = re.sub(r"[^\d,.\-]", "", str(val).replace("\xa0", "").strip())
    if not cleaned:
        return None

    sign = -1 if cleaned.count("-") % 2 else 1
    cleaned = cleaned.replace("-", "")
    if not cleaned:
        return None

    last_dot = cleaned.rfind(".")
    last_comma = cleaned.rfind(",")

    decimal_sep = None
    if last_dot != -1 and last_comma != -1:
        decimal_sep = "." if last_dot > last_comma else ","
    elif last_comma != -1 and len(cleaned) - last_comma - 1 in (1, 2):
        decimal_sep = ","
    elif last_dot != -1 and len(cleaned) - last_dot - 1 in (1, 2):
        decimal_sep = "."

    if decimal_sep:
        thousands_sep = "," if decimal_sep == "." else "."
        cleaned = cleaned.replace(thousands_sep, "")
        integer_part, fractional_part = cleaned.rsplit(decimal_sep, 1)
        normalized = f"{integer_part}.{fractional_part}"
    else:
        normalized = cleaned.replace(",", "").replace(".", "")

    try:
        return round(sign * float(normalized), 2)
    except ValueError:
        return None


def parse_date(val: Optional[str]) -> Optional[str]:
    if not val:
        return None
    m = re.match(r"(\d{2})\.(\d{2})\.(\d{4})\s+(\d{2}:\d{2}:\d{2})", val)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}T{m.group(4)}"
    return val


def parse_margin(val: Optional[str]) -> Optional[float]:
    if not val:
        return None
    cleaned = val.replace("%", "").replace(",", ".").strip()
    try:
        return round(float(cleaned), 1)
    except ValueError:
        return None


def normalize_lieferant_name(val: Optional[str]) -> Optional[str]:
    if not val:
        return None

    normalized = re.sub(r"\s+", " ", str(val)).strip()
    if not normalized:
        return None

    return LIEFERANT_ALIASES.get(normalized.lower(), normalized)


def make_placeholder_product(sku: str, lieferanten_map: Dict[str, str]) -> Dict[str, Any]:
    product = {column: None for column in PRODUCT_COLUMNS}
    product["sku"] = sku
    product["lieferant"] = lieferanten_map.get(sku)
    return product


def parse_sales():
    tree = ET.parse(SALES_XML)
    root = tree.getroot()
    worksheet = root.findall(".//ss:Worksheet", NS)[0]
    table = worksheet.find("ss:Table", NS)
    rows = table.findall("ss:Row", NS)

    headers = []
    for cell in rows[0].findall("ss:Cell", NS):
        data = cell.find("ss:Data", NS)
        headers.append(data.text if data is not None else "")

    sales = []
    filtered_90 = 0
    skipped_summary = 0
    for row in rows[1:]:
        cells = row.findall("ss:Cell", NS)
        values = []
        col_idx = 0
        for cell in cells:
            idx_attr = cell.attrib.get(f"{{{NS['ss']}}}Index")
            if idx_attr:
                target = int(idx_attr) - 1
                while col_idx < target:
                    values.append(None)
                    col_idx += 1
            data = cell.find("ss:Data", NS)
            values.append(data.text if data is not None else None)
            col_idx += 1

        raw = dict(zip(headers, values))
        order_no = (raw.get("Bestellung #") or "").strip()
        order_date = (raw.get("Bestelldatum") or "").strip()
        art = (raw.get("Artikelposition") or "").strip()
        if not order_no and not order_date and not art:
            skipped_summary += 1
            continue

        if art and art.startswith("90"):
            filtered_90 += 1
            continue

        record = {}
        for orig_key, camel_key in SALES_KEY_MAP.items():
            val = raw.get(orig_key)
            if camel_key in MONEY_FIELDS:
                record[camel_key] = parse_money(val)
            elif camel_key in QTY_FIELDS:
                record[camel_key] = int(float(val)) if val else 0
            elif camel_key == "bestelldatum":
                record[camel_key] = parse_date(val)
            elif camel_key == "totalMargin":
                record[camel_key] = parse_margin(val)
            else:
                record[camel_key] = val
        sales.append(record)

    print(f"Sales: {len(sales)} rows loaded, {filtered_90} filtered (90*), {skipped_summary} summary rows skipped")
    return sales


def parse_lieferanten() -> Dict[str, str]:
    """Загружает SKU → Lieferant маппинг из отдельного файла."""
    if not LIEFERANT_XLSX.exists():
        print(f"Lieferanten file not found: {LIEFERANT_XLSX}")
        return {}

    wb = openpyxl.load_workbook(LIEFERANT_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        wb.close()
        return {}

    sku_idx = 0
    lf_idx = 1
    for i, name in enumerate(header):
        if name and str(name).strip().lower() == "sku":
            sku_idx = i
        if name and str(name).strip().lower() in ("lieferant", "supplier", "поставщик"):
            lf_idx = i

    mapping: Dict[str, str] = {}
    for row in rows_iter:
        if not row:
            continue
        sku = str(row[sku_idx]).strip() if row[sku_idx] is not None else ""
        lf = normalize_lieferant_name(row[lf_idx] if len(row) > lf_idx else None) or ""
        if sku and lf:
            mapping[sku] = lf
    wb.close()
    print(f"Lieferanten: {len(mapping)} SKU → поставщик")
    return mapping


def parse_catalog(sales_skus: Set[str], inventory_skus: Set[str], lieferanten_map: Dict[str, str]):
    wb = openpyxl.load_workbook(CATALOG_XLSX, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=False)
    header_row = next(rows_iter)
    headers = [cell.value for cell in header_row]
    relevant_skus = sales_skus | inventory_skus

    col_indices = {}
    for col_name in PRODUCT_COLUMNS:
        if col_name in headers:
            col_indices[col_name] = headers.index(col_name)

    products = {}
    parent_groups = {}
    lieferanten = set()
    total = 0
    matched = 0

    for row in rows_iter:
        total += 1
        sku_cell = row[col_indices["sku"]]
        sku = str(sku_cell.value).strip() if sku_cell.value else None
        if not sku:
            continue

        parent_sku = None
        if "amaz_parent_sku" in col_indices:
            pval = row[col_indices["amaz_parent_sku"]].value
            parent_sku = str(pval).strip() if pval else None

        if parent_sku:
            parent_groups.setdefault(parent_sku, [])
            if sku not in parent_groups[parent_sku]:
                parent_groups[parent_sku].append(sku)

        if sku not in relevant_skus:
            if parent_sku:
                has_relevant_sibling = any(s in relevant_skus for s in parent_groups.get(parent_sku, []))
                if not has_relevant_sibling:
                    continue
            else:
                continue

        matched += 1
        product = {}
        for col_name, idx in col_indices.items():
            val = row[idx].value
            if val is not None:
                if col_name in ("purchase_price", "price", "amaz_price"):
                    try:
                        product[col_name] = round(float(val), 2)
                    except (ValueError, TypeError):
                        product[col_name] = str(val)
                else:
                    product[col_name] = str(val).strip()
            else:
                product[col_name] = None
        product["lieferant"] = lieferanten_map.get(sku)
        products[sku] = product

        if product["lieferant"]:
            lieferanten.add(product["lieferant"])

    wb.close()

    # Second pass: include siblings of matched SKUs
    sibling_skus = set()
    for parent, children in parent_groups.items():
        if any(c in relevant_skus for c in children):
            for c in children:
                if c not in products:
                    sibling_skus.add(c)

    if sibling_skus:
        wb = openpyxl.load_workbook(CATALOG_XLSX, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=False)
        next(rows_iter)  # skip header
        for row in rows_iter:
            sku_cell = row[col_indices["sku"]]
            sku = str(sku_cell.value).strip() if sku_cell.value else None
            if sku not in sibling_skus:
                continue
            product = {}
            for col_name, idx in col_indices.items():
                val = row[idx].value
                if val is not None:
                    if col_name in ("purchase_price", "price", "amaz_price"):
                        try:
                            product[col_name] = round(float(val), 2)
                        except (ValueError, TypeError):
                            product[col_name] = str(val)
                    else:
                        product[col_name] = str(val).strip()
                else:
                    product[col_name] = None
            product["lieferant"] = lieferanten_map.get(sku)
            if product["lieferant"]:
                lieferanten.add(product["lieferant"])
            products[sku] = product
        wb.close()

    missing_relevant = relevant_skus - set(products.keys())
    for sku in sorted(missing_relevant):
        product = make_placeholder_product(sku, lieferanten_map)
        products[sku] = product
        if product["lieferant"]:
            lieferanten.add(product["lieferant"])

    # Filter parent_groups to only those with sales/inventory-relevant children
    relevant_parents = {}
    for parent, children in parent_groups.items():
        if any(c in relevant_skus for c in children):
            relevant_parents[parent] = children

    unmatched = sales_skus - set(products.keys())
    inventory_only = inventory_skus - sales_skus
    matched_inventory = inventory_skus & set(products.keys())
    print(f"Catalog: {total} total rows, {len(products)} products exported")
    print(f"Parent groups: {len(relevant_parents)}")
    print(f"Lieferanten: {sorted(lieferanten)}")
    print(f"Inventory SKUs in catalog: {len(matched_inventory)}/{len(inventory_skus)}")
    print(f"Inventory-only SKUs considered: {len(inventory_only)}")
    print(f"Placeholder products added: {len(missing_relevant)}")
    if unmatched:
        print(f"Unmatched sales SKUs: {sorted(unmatched)}")

    return {
        "products": products,
        "parentGroups": relevant_parents,
        "lieferanten": sorted(lieferanten),
    }


def parse_inventory():
    records = {}

    with open(INVENTORY_TXT, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            sku = (row.get("seller-sku") or "").strip()
            if not sku:
                continue

            quantity = int(float(row.get("Quantity Available") or 0))
            condition = (row.get("Warehouse-Condition-code") or "").strip().upper()

            record = records.setdefault(sku, {
                "sku": sku,
                "asin": (row.get("asin") or "").strip() or None,
                "fulfillmentChannelSku": (row.get("fulfillment-channel-sku") or "").strip() or None,
                "sellable": 0,
                "unsellable": 0,
                "total": 0,
            })

            if condition == "SELLABLE":
                record["sellable"] += quantity
            else:
                record["unsellable"] += quantity

            record["total"] = record["sellable"] + record["unsellable"]

    totals = {
        "sellable": sum(item["sellable"] for item in records.values()),
        "unsellable": sum(item["unsellable"] for item in records.values()),
        "total": sum(item["total"] for item in records.values()),
        "skusWithStock": sum(1 for item in records.values() if item["sellable"] > 0),
        "trackedSkus": len(records),
    }

    print(f"Inventory: {len(records)} SKUs, {totals['sellable']} sellable units")

    return {
        "records": records,
        "totals": totals,
    }


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    sales = parse_sales()
    sales_skus = {r["artikelposition"] for r in sales if r.get("artikelposition")}
    inventory = parse_inventory()
    inventory_skus = set(inventory["records"].keys())

    lieferanten_map = parse_lieferanten()
    catalog = parse_catalog(sales_skus, inventory_skus, lieferanten_map)

    with open(OUT_DIR / "sales.json", "w", encoding="utf-8") as f:
        json.dump(sales, f, ensure_ascii=False, indent=2)

    with open(OUT_DIR / "products.json", "w", encoding="utf-8") as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)

    with open(OUT_DIR / "inventory.json", "w", encoding="utf-8") as f:
        json.dump(inventory, f, ensure_ascii=False, indent=2)

    print(f"\nOutput: {OUT_DIR / 'sales.json'}, {OUT_DIR / 'products.json'}, {OUT_DIR / 'inventory.json'}")


if __name__ == "__main__":
    main()
