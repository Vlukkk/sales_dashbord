#!/usr/bin/env python3
"""Build an Excel audit report joining SKU-Lieferant with Amazon inventory."""

from __future__ import annotations

from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from preprocess import (
        BASE_DIR,
        INVENTORY_TXT,
        LIEFERANT_XLSX,
        parse_inventory,
        parse_lieferanten,
    )
except ModuleNotFoundError:
    from scripts.preprocess import (
        BASE_DIR,
        INVENTORY_TXT,
        LIEFERANT_XLSX,
        parse_inventory,
        parse_lieferanten,
    )

REPORTS_DIR = BASE_DIR / "reports"
OUTPUT_PATH = REPORTS_DIR / f"sku_lieferant_inventory_audit_{date.today().isoformat()}.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="EAF2FF")
META_FILL = PatternFill("solid", fgColor="F8FAFC")
ALERT_FILL = PatternFill("solid", fgColor="FEF2F2")
SUBTLE_FILL = PatternFill("solid", fgColor="F8FAFC")
HEADER_FONT = Font(bold=True, color="0F172A")
LINK_FONT = Font(color="2563EB")


def autosize_columns(ws) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for index, value in enumerate(row, start=1):
            if value is None:
                continue
            widths[index] = max(widths.get(index, 0), len(str(value)))

    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 12), 42)


def style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def apply_filter_and_freeze(ws) -> None:
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def build_joined_rows(
    supplier_map: Dict[str, str],
    inventory_records: Dict[str, dict],
) -> List[dict]:
    rows: List[dict] = []
    for sku, record in inventory_records.items():
        supplier = supplier_map.get(sku)
        rows.append(
            {
                "lieferant": supplier or "Без поставщика",
                "sku": sku,
                "matchedSupplier": "Да" if supplier else "Нет",
                "asin": record.get("asin"),
                "fulfillmentChannelSku": record.get("fulfillmentChannelSku"),
                "sellable": record.get("sellable", 0),
                "unsellable": record.get("unsellable", 0),
                "total": record.get("total", 0),
            }
        )

    rows.sort(key=lambda item: (item["lieferant"] == "Без поставщика", item["lieferant"], item["sku"]))
    return rows


def build_summary_rows(joined_rows: List[dict]) -> List[dict]:
    grouped = defaultdict(
        lambda: {
            "skus": set(),
            "sellable": 0,
            "unsellable": 0,
            "total": 0,
            "skusWithStock": 0,
        }
    )

    for row in joined_rows:
        group = grouped[row["lieferant"]]
        group["skus"].add(row["sku"])
        group["sellable"] += row["sellable"]
        group["unsellable"] += row["unsellable"]
        group["total"] += row["total"]
        if row["sellable"] > 0:
            group["skusWithStock"] += 1

    summary = [
        {
            "lieferant": supplier,
            "skuCount": len(values["skus"]),
            "skusWithStock": values["skusWithStock"],
            "sellable": values["sellable"],
            "unsellable": values["unsellable"],
            "total": values["total"],
        }
        for supplier, values in grouped.items()
    ]
    summary.sort(key=lambda item: (-item["total"], -item["sellable"], item["lieferant"]))
    return summary


def write_overview_sheet(
    wb: Workbook,
    supplier_map: Dict[str, str],
    inventory_records: Dict[str, dict],
    joined_rows: List[dict],
    summary_rows: List[dict],
) -> None:
    ws = wb.active
    ws.title = "Overview"

    matched_count = sum(1 for row in joined_rows if row["matchedSupplier"] == "Да")
    inventory_only = len(joined_rows) - matched_count
    supplier_only = len([sku for sku in supplier_map if sku not in inventory_records])

    ws["A1"] = "SKU-Lieferant vs Amazon Inventory Audit"
    ws["A1"].font = Font(size=14, bold=True, color="0F172A")
    ws["A3"] = "Source file"
    ws["B3"] = str(LIEFERANT_XLSX.name)
    ws["A4"] = "Inventory file"
    ws["B4"] = str(INVENTORY_TXT.name)
    ws["A5"] = "Generated"
    ws["B5"] = date.today().isoformat()

    for row in range(3, 6):
        ws[f"A{row}"].fill = META_FILL
        ws[f"A{row}"].font = HEADER_FONT

    metrics = [
        ("SKU in supplier file", len(supplier_map)),
        ("SKU in inventory file", len(inventory_records)),
        ("Inventory SKU matched to supplier", matched_count),
        ("Inventory SKU without supplier", inventory_only),
        ("Supplier SKU not in inventory", supplier_only),
        ("Suppliers in summary", len(summary_rows)),
        ("Sellable units total", sum(row["sellable"] for row in joined_rows)),
        ("Unsellable units total", sum(row["unsellable"] for row in joined_rows)),
    ]

    ws["A8"] = "Key metrics"
    ws["A8"].font = Font(size=12, bold=True, color="0F172A")
    for index, (label, value) in enumerate(metrics, start=9):
        ws[f"A{index}"] = label
        ws[f"B{index}"] = value
        ws[f"A{index}"].fill = META_FILL

    ws["D8"] = "How to read"
    ws["D8"].font = Font(size=12, bold=True, color="0F172A")
    notes = [
        "Join key: SKU-Lieferant.sku = inventory.seller-sku",
        "Lieferant names are normalized with the same aliases as preprocess.py",
        "Sheet 'Summary by Lieferant' is the main control pivot for stock by supplier",
        "Sheet 'Inventory without Lieferant' shows SKUs present in Amazon stock but missing in SKU-Lieferant.xlsx",
    ]
    for index, note in enumerate(notes, start=9):
        ws[f"D{index}"] = note

    ws["A19"] = "Top suppliers by total units"
    ws["A19"].font = Font(size=12, bold=True, color="0F172A")
    headers = ["Lieferant", "SKU", "SKU with stock", "Sellable", "Unsellable", "Total"]
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row=20, column=column, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_index, row in enumerate(summary_rows[:15], start=21):
        ws.cell(row=row_index, column=1, value=row["lieferant"])
        ws.cell(row=row_index, column=2, value=row["skuCount"])
        ws.cell(row=row_index, column=3, value=row["skusWithStock"])
        ws.cell(row=row_index, column=4, value=row["sellable"])
        ws.cell(row=row_index, column=5, value=row["unsellable"])
        ws.cell(row=row_index, column=6, value=row["total"])

    autosize_columns(ws)


def write_summary_sheet(wb: Workbook, summary_rows: List[dict]) -> None:
    ws = wb.create_sheet("Summary by Lieferant")
    headers = ["Lieferant", "SKU count", "SKU with stock", "Sellable", "Unsellable", "Total"]
    ws.append(headers)
    for row in summary_rows:
        ws.append(
            [
                row["lieferant"],
                row["skuCount"],
                row["skusWithStock"],
                row["sellable"],
                row["unsellable"],
                row["total"],
            ]
        )

    style_header(ws)
    apply_filter_and_freeze(ws)
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "Без поставщика":
            for cell in row:
                cell.fill = ALERT_FILL
    autosize_columns(ws)


def write_joined_sheet(wb: Workbook, joined_rows: List[dict]) -> None:
    ws = wb.create_sheet("Joined inventory")
    headers = ["Lieferant", "SKU", "Matched supplier", "ASIN", "FNSKU", "Sellable", "Unsellable", "Total"]
    ws.append(headers)
    for row in joined_rows:
        ws.append(
            [
                row["lieferant"],
                row["sku"],
                row["matchedSupplier"],
                row["asin"],
                row["fulfillmentChannelSku"],
                row["sellable"],
                row["unsellable"],
                row["total"],
            ]
        )

    style_header(ws)
    apply_filter_and_freeze(ws)
    for row in ws.iter_rows(min_row=2):
        if row[2].value == "Нет":
            for cell in row:
                cell.fill = ALERT_FILL
    autosize_columns(ws)


def write_unmatched_inventory_sheet(wb: Workbook, joined_rows: List[dict]) -> None:
    ws = wb.create_sheet("Inventory without Lieferant")
    headers = ["SKU", "ASIN", "FNSKU", "Sellable", "Unsellable", "Total"]
    ws.append(headers)
    for row in joined_rows:
        if row["matchedSupplier"] == "Да":
            continue
        ws.append(
            [
                row["sku"],
                row["asin"],
                row["fulfillmentChannelSku"],
                row["sellable"],
                row["unsellable"],
                row["total"],
            ]
        )

    style_header(ws)
    apply_filter_and_freeze(ws)
    autosize_columns(ws)


def write_supplier_only_sheet(wb: Workbook, supplier_map: Dict[str, str], inventory_records: Dict[str, dict]) -> None:
    ws = wb.create_sheet("Lieferant without inventory")
    headers = ["SKU", "Lieferant"]
    ws.append(headers)
    for sku, supplier in sorted(supplier_map.items(), key=lambda item: (item[1], item[0])):
        if sku in inventory_records:
            continue
        ws.append([sku, supplier])

    style_header(ws)
    apply_filter_and_freeze(ws)
    autosize_columns(ws)


def main() -> None:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    supplier_map = parse_lieferanten()
    inventory = parse_inventory()
    inventory_records = inventory["records"]

    joined_rows = build_joined_rows(supplier_map, inventory_records)
    summary_rows = build_summary_rows(joined_rows)

    wb = Workbook()
    write_overview_sheet(wb, supplier_map, inventory_records, joined_rows, summary_rows)
    write_summary_sheet(wb, summary_rows)
    write_joined_sheet(wb, joined_rows)
    write_unmatched_inventory_sheet(wb, joined_rows)
    write_supplier_only_sheet(wb, supplier_map, inventory_records)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

    wb.save(OUTPUT_PATH)
    print(f"Audit report saved to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
